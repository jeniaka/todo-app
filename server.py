import json
import os
import re
import hmac
import hashlib
import base64
import secrets
import mimetypes
import time
import queue
import threading
import urllib.request
import urllib.parse
import uuid as _uuid
import csv
import io
from datetime import datetime
from http.server import HTTPServer, ThreadingHTTPServer, BaseHTTPRequestHandler
from http.cookies import SimpleCookie
from urllib.parse import urlparse, parse_qs
from mail import (
    smtp_configured, send_email, smtp_status, print_smtp_status,
    build_email, build_invite_email, build_task_assigned_email,
    build_task_done_email, build_overdue_email, build_member_joined_email,
    SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM, SMTP_DEBUG,
)

# ─── Config ──────────────────────────────────────────────────────────────────
PORT               = int(os.environ.get("PORT", 8090))
MONGODB_URI        = os.environ.get("MONGODB_URI", "")
GOOGLE_CLIENT_ID   = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
REDIRECT_URI       = os.environ.get("REDIRECT_URI", f"http://localhost:{PORT}/auth/callback")
SESSION_SECRET     = os.environ.get("SESSION_SECRET", secrets.token_hex(32))
ANTHROPIC_API_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")

def check_and_send_overdue_emails(user):
    """Check for overdue tasks and send notification emails."""
    if not smtp_configured():
        return
    now = int(time.time() * 1000)

    # Personal tasks
    todos = load_todos(user["id"])
    changed = False
    for task in todos:
        if (task.get("dueDate") and
                task["dueDate"] < now and
                not task.get("done") and
                task.get("status", "todo") != "done" and
                not task.get("overdueNotified")):
            html = build_overdue_email(task["text"], task.get("dueDate"), None)
            _ok, _err = send_email(user.get("email", ""), f'⚠ Overdue: "{task["text"]}"', html)
            if not _ok:
                print(f"[overdue-email] failed for task={task['text']!r}: {_err}")
            task["overdueNotified"] = True
            changed = True
    if changed:
        save_todos(user["id"], todos)

    # Group tasks
    if db is None:
        return
    groups = load_groups_for_user(user["id"])
    for g in groups:
        g_changed = False
        for task in g.get("tasks", []):
            if (task.get("dueDate") and
                    task["dueDate"] < now and
                    not task.get("done") and
                    task.get("status", "todo") != "done" and
                    not task.get("overdueNotified") and
                    task.get("assignedTo") == user["id"]):
                assignee_email = user.get("email", "")
                html = build_overdue_email(task["text"], task.get("dueDate"), g["name"])
                _ok, _err = send_email(assignee_email, f'⚠ Overdue: "{task["text"]}" in {g["name"]}', html)
                if not _ok:
                    print(f"[overdue-email] failed for task={task['text']!r}: {_err}")
                task["overdueNotified"] = True
                g_changed = True
        if g_changed:
            save_group(g)


BASE = os.path.dirname(__file__)
STATIC_DIR    = os.path.join(BASE, "static")
TEMPLATES_DIR = os.path.join(BASE, "templates")
DATA_DIR      = os.path.join(BASE, "userdata")
os.makedirs(DATA_DIR, exist_ok=True)

# ─── Default settings ─────────────────────────────────────────────────────────
DEFAULT_SETTINGS = {
    "language": "en",
    "notifications": {
        "taskReminders": True,
        "dailyDigest": False,
        "taskCompleted": True,
        "overdueTasks": True,
        "weeklyReport": False,
        "notificationSound": True,
        "badgeCount": True,
        "doNotDisturb": False
    }
}

# ─── MongoDB ──────────────────────────────────────────────────────────────────
db = None
DB_STATUS = "local_files"
if MONGODB_URI:
    try:
        from pymongo import MongoClient
        import certifi
        client = MongoClient(MONGODB_URI, tlsCAFile=certifi.where(), serverSelectionTimeoutMS=5000)
        client.admin.command('ping')
        db = client["tododb"]
        DB_STATUS = "mongodb"
        print("✓ MongoDB connected successfully")
    except Exception as e:
        print(f"✗ MongoDB connection failed: {e}")
        db = None
else:
    pass

if DB_STATUS == "local_files":
    print("╔══════════════════════════════════════════════════════════╗")
    print("║  ⚠ WARNING: MongoDB not connected!                      ║")
    print("║  Using local file storage — DATA WILL BE LOST ON DEPLOY  ║")
    print("║  Set MONGODB_URI environment variable to fix this.       ║")
    print("╚══════════════════════════════════════════════════════════╝")

def load_todos(user_id):
    if db is not None:
        doc = db["todos"].find_one({"_id": user_id})
        return doc["data"] if doc else []
    path = os.path.join(DATA_DIR, f"{user_id}.json")
    return json.load(open(path)) if os.path.exists(path) else []

def save_todos(user_id, data):
    if db is not None:
        db["todos"].replace_one({"_id": user_id}, {"_id": user_id, "data": data}, upsert=True)
        return
    with open(os.path.join(DATA_DIR, f"{user_id}.json"), "w") as f:
        json.dump(data, f, indent=2)

def load_settings(user_id):
    if db is not None:
        doc = db["settings"].find_one({"_id": user_id})
        return doc["data"] if doc else DEFAULT_SETTINGS
    path = os.path.join(DATA_DIR, f"{user_id}_settings.json")
    return json.load(open(path)) if os.path.exists(path) else DEFAULT_SETTINGS

def save_settings(user_id, data):
    if db is not None:
        db["settings"].replace_one({"_id": user_id}, {"_id": user_id, "data": data}, upsert=True)
        return
    with open(os.path.join(DATA_DIR, f"{user_id}_settings.json"), "w") as f:
        json.dump(data, f, indent=2)

# ─── Groups helpers ───────────────────────────────────────────────────────────
def new_id():
    return str(_uuid.uuid4())

def generate_slug(name, existing_slugs=None):
    """Generate a URL-friendly slug from a group name."""
    s = name.lower().strip()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = s.strip('-')[:50]
    if not s:
        s = secrets.token_hex(4)
    if existing_slugs is None:
        existing_slugs = []
    base = s
    counter = 2
    while s in existing_slugs:
        s = f"{base}-{counter}"
        counter += 1
    return s

def migrate_group_slugs():
    """Add slugs to groups that don't have them (idempotent)."""
    if db is None:
        return
    groups_without = list(db["groups"].find({"slug": {"$exists": False}}))
    if not groups_without:
        return
    existing = [g.get("slug","") for g in db["groups"].find({"slug": {"$exists": True}}, {"slug": 1})]
    for g in groups_without:
        slug = generate_slug(g.get("name","group"), existing)
        db["groups"].update_one({"_id": g["_id"]}, {"$set": {"slug": slug}})
        existing.append(slug)
        print(f"  Migrated group '{g.get('name')}' → slug: {slug}")

def load_groups_for_user(user_id):
    """Return all groups where user is an active member."""
    if db is None: return []
    return list(db["groups"].find({"members": {"$elemMatch": {"userId": user_id, "status": "active"}}}))

def load_group(group_id):
    if db is None: return None
    return db["groups"].find_one({"_id": group_id})

# ─── SSE / Real-time push ─────────────────────────────────────────────────────
_sse_clients: dict = {}   # group_id -> list[queue.Queue]  (pre-encoded SSE frames)
_sse_lock = threading.Lock()

def sse_broadcast(group_id, by_user_id):
    """Push an update event to every SSE client watching this group."""
    msg = (b'data: ' +
           json.dumps({"type": "update", "groupId": str(group_id), "byUserId": by_user_id}).encode() +
           b'\n\n')
    with _sse_lock:
        clients = list(_sse_clients.get(str(group_id), []))
    dead = []
    for q in clients:
        try:
            q.put_nowait(msg)
        except Exception:
            dead.append(q)
    if dead:
        with _sse_lock:
            bucket = _sse_clients.get(str(group_id), [])
            for q in dead:
                try: bucket.remove(q)
                except ValueError: pass

def save_group(group, changed_by=None):
    if db is None: return
    db["groups"].replace_one({"_id": group["_id"]}, group, upsert=True)
    if changed_by is not None:
        sse_broadcast(group["_id"], changed_by)

def get_member_role(group, user_id):
    for m in group.get("members", []):
        if m.get("userId") == user_id and m.get("status") == "active":
            return m.get("role")
    return None

def create_notification(user_id, notif_type, group_id, group_name, task_text=None, from_user=None):
    if db is None: return
    if not user_id: return
    db["notifications"].insert_one({
        "_id": new_id(),
        "userId": user_id,
        "type": notif_type,
        "groupId": group_id,
        "groupName": group_name,
        "taskText": task_text,
        "fromUser": from_user,
        "createdAt": int(time.time() * 1000),
        "read": False,
    })

def record_activity(group_id, activity):
    if db is None: return
    try:
        db["groups"].update_one(
            {"_id": group_id},
            {"$push": {"activities": {"$each": [activity], "$slice": -30}}}
        )
    except Exception as e:
        print(f"⚠ record_activity failed: {e}")

# ─── Anthropic AI ─────────────────────────────────────────────────────────────
def ai_suggest_next(todos):
    if not ANTHROPIC_API_KEY:
        return {"result": "AI is not configured. Add ANTHROPIC_API_KEY to your environment variables."}
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    open_tasks = [t for t in todos if not t.get("done")]
    if not open_tasks:
        return {"result": "All tasks are complete — great work! Add new tasks to get a suggestion."}
    task_list = "\n".join(f"- [{t.get('priority','none')}] {t['text']} (created {t.get('createdAt',0)})" for t in open_tasks[:20])
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=256,
        messages=[{"role": "user", "content": f"Given these open tasks, which one should I tackle next and why? Be concise (2-3 sentences).\n\n{task_list}"}]
    )
    return {"result": message.content[0].text}

def ai_daily_summary(todos):
    if not ANTHROPIC_API_KEY:
        return {"result": "AI is not configured. Add ANTHROPIC_API_KEY to your environment variables."}
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    done  = [t for t in todos if t.get("done")]
    open_ = [t for t in todos if not t.get("done")]
    summary_input = f"Done today: {len(done)} tasks. Still open: {len(open_)} tasks.\nCompleted: {', '.join(t['text'] for t in done[:10])}\nOpen: {', '.join(t['text'] for t in open_[:10])}"
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=200,
        messages=[{"role": "user", "content": f"Give a brief, encouraging daily summary for this task list. Keep it to 2-3 sentences.\n\n{summary_input}"}]
    )
    return {"result": message.content[0].text}

def ai_split_task(title, description):
    if not ANTHROPIC_API_KEY:
        return {"subtasks": ["Set up the project", "Research the topic", "Draft the outline", "Review and finalize"]}
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = f"Break this task into 3-6 specific, actionable sub-tasks. Return ONLY a JSON array of strings, nothing else.\n\nTask: {title}"
    if description:
        prompt += f"\nDescription: {description}"
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}]
    )
    text = message.content[0].text.strip()
    try:
        start = text.index('[')
        end   = text.rindex(']') + 1
        subtasks = json.loads(text[start:end])
        return {"subtasks": subtasks}
    except Exception:
        lines = [l.strip().lstrip('-•*123456789. ') for l in text.split('\n') if l.strip()]
        return {"subtasks": [l for l in lines if l][:6]}

# ─── Session ──────────────────────────────────────────────────────────────────
def sign(data):
    return hmac.new(SESSION_SECRET.encode(), data.encode(), hashlib.sha256).hexdigest()

def create_session(user):
    payload = base64.urlsafe_b64encode(json.dumps(user).encode()).decode()
    return f"{payload}.{sign(payload)}"

def verify_session(token):
    try:
        payload, sig = token.rsplit(".", 1)
        if hmac.compare_digest(sign(payload), sig):
            return json.loads(base64.urlsafe_b64decode(payload))
    except Exception:
        pass
    return None

def create_state():
    s = secrets.token_urlsafe(16)
    return f"{s}.{sign(s)[:16]}"

def verify_state(state_tok):
    try:
        s, sig = state_tok.rsplit(".", 1)
        return hmac.compare_digest(sign(s)[:16], sig)
    except Exception:
        return False

# ─── Google OAuth ─────────────────────────────────────────────────────────────
def request_redirect_uri(handler):
    """Build redirect_uri from the incoming request's Host header so the app
    works on any domain (mytasks.bar, onrender.com, localhost) without needing
    to update the REDIRECT_URI env var each time the domain changes."""
    host = handler.headers.get("Host", "")
    if not host:
        return REDIRECT_URI  # fallback to env var
    is_local = host.startswith("localhost") or host.startswith("127.")
    scheme = "http" if is_local else "https"
    return f"{scheme}://{host}/auth/callback"

def google_auth_url(state_tok, redirect_uri):
    params = urllib.parse.urlencode({
        "client_id": GOOGLE_CLIENT_ID, "redirect_uri": redirect_uri,
        "response_type": "code", "scope": "openid email profile",
        "state": state_tok, "prompt": "select_account",
    })
    return f"https://accounts.google.com/o/oauth2/v2/auth?{params}"

def exchange_code(code, redirect_uri):
    data = urllib.parse.urlencode({
        "code": code, "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET, "redirect_uri": redirect_uri,
        "grant_type": "authorization_code",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data)
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def get_google_user(access_token):
    req = urllib.request.Request(
        "https://www.googleapis.com/oauth2/v3/userinfo",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

# ─── Templates ────────────────────────────────────────────────────────────────
def load_template(name):
    path = os.path.join(TEMPLATES_DIR, name)
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def render_app(user):
    name    = user.get("name", "")
    email   = user.get("email", "")
    picture = user.get("picture", "")
    avatar  = (f'<img class="user-avatar" src="{picture}" referrerpolicy="no-referrer" alt="{name}">'
               if picture else
               f'<div class="user-avatar-fallback">{name[:1].upper()}</div>')
    user_json = json.dumps({"name": name, "email": email, "picture": picture, "id": user.get("id", "")})
    html = load_template("app.html")
    html = (html.replace("{{AVATAR}}", avatar)
                .replace("{{NAME}}", name)
                .replace("{{EMAIL}}", email)
                .replace("{{USER_JSON}}", user_json))
    return html.encode()

# ─── Static assets ────────────────────────────────────────────────────────────
MANIFEST = json.dumps({
    "name": "My Tasks", "short_name": "Tasks",
    "start_url": "/", "display": "standalone",
    "background_color": "#F8F7F4", "theme_color": "#F8F7F4",
    "icons": [
        {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
        {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
    ]
}).encode()

SERVICE_WORKER = b"""
const CACHE = 'tasks-v3';
self.addEventListener('install', e => e.waitUntil(caches.open(CACHE).then(c => c.addAll(['/login', '/static/style.css', '/static/app.js']))));
self.addEventListener('fetch', e => {
  if (e.request.url.includes('/api/') || e.request.url.includes('/auth/')) return;
  e.respondWith(fetch(e.request).catch(() => caches.match(e.request)));
});
"""

# ─── Import / CSV helpers ─────────────────────────────────────────────────────
def parse_csv_rows(file_content):
    """Parse CSV bytes/str and return list of dicts."""
    if isinstance(file_content, bytes):
        file_content = file_content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(file_content))
    return [row for row in reader if any(v.strip() for v in row.values())]

def parse_xlsx_rows(file_bytes):
    """Parse XLSX bytes and return list of dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                result.append({headers[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(headers))})
        wb.close()
        return result
    except Exception as e:
        print(f"[xlsx] parse error: {e}")
        return []

def detect_platform(headers):
    """Detect whether rows came from Jira, Monday.com, or a generic CSV."""
    hl = [h.lower().strip() for h in headers]
    jira_fields = {"issue key", "issue id", "summary", "issue type", "story points", "sprint"}
    if jira_fields & set(hl):
        return "jira"
    monday_fields = {"subitems", "board", "pulse id", "item id"}
    if monday_fields & set(hl):
        return "monday"
    return "generic"

def map_row_to_task(row, platform):
    """Map one CSV/XLSX row to a MyTasks task dict."""
    norm = {k.lower().strip(): v for k, v in row.items()}
    now_ms = int(time.time() * 1000)
    task = {
        "id": secrets.token_hex(8),
        "text": "",
        "description": "",
        "status": "todo",
        "priority": "none",
        "assignedTo": None,
        "assigneeEmail": None,
        "assigneeName": None,
        "createdAt": now_ms,
        "completedAt": None,
        "startedAt": None,
        "estimatedHours": None,
        "dueDate": None,
        "subtasks": [],
        "done": False,
        "importedFrom": platform,
    }
    # Title
    for key in ["summary", "title", "task", "name", "item", "issue", "task name", "item name"]:
        if key in norm and norm[key].strip():
            task["text"] = norm[key].strip()[:500]
            break
    if not task["text"]:
        for v in row.values():
            if v and str(v).strip():
                task["text"] = str(v).strip()[:500]
                break
    if not task["text"]:
        return None  # skip empty rows
    # Description
    for key in ["description", "details", "notes", "comment", "description/steps to reproduce"]:
        if key in norm and norm[key].strip():
            task["description"] = norm[key].strip()[:2000]
            break
    # Status
    for key in ["status", "state", "stage", "workflow"]:
        if key in norm and norm[key].strip():
            rs = norm[key].strip().lower()
            if rs in {"done", "complete", "completed", "closed", "resolved", "finished"}:
                task["status"] = "done"
                task["done"] = True
                task["completedAt"] = now_ms
            elif rs in {"in progress", "in review", "working on it", "active", "started", "doing", "in development"}:
                task["status"] = "in_progress"
                task["startedAt"] = now_ms
            else:
                task["status"] = "todo"
            break
    # Priority
    for key in ["priority", "urgency", "importance"]:
        if key in norm and norm[key].strip():
            rp = norm[key].strip().lower()
            if rp in {"highest", "critical", "urgent", "high", "blocker"}:
                task["priority"] = "high"
            elif rp in {"medium", "normal", "moderate", "mid", "medium/normal"}:
                task["priority"] = "medium"
            elif rp in {"low", "minor", "trivial", "lowest"}:
                task["priority"] = "low"
            break
    # Assignee
    for key in ["assignee", "assigned to", "owner", "person", "people", "responsible"]:
        if key in norm and norm[key].strip():
            raw = norm[key].strip()
            if "@" in raw:
                task["assigneeEmail"] = raw.lower()
            else:
                task["assigneeName"] = raw
            break
    # Due date
    for key in ["due date", "due", "deadline", "target date", "end date"]:
        if key in norm and norm[key].strip():
            ds = norm[key].strip()
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y", "%B %d, %Y"]:
                try:
                    dt = datetime.strptime(ds.split(".")[0].split("+")[0].strip(), fmt)
                    task["dueDate"] = int(dt.timestamp() * 1000)
                    break
                except ValueError:
                    continue
            break
    # Estimated hours / story points
    for key in ["story points", "estimate", "estimated hours", "time estimate", "original estimate"]:
        if key in norm and norm[key].strip():
            try:
                val = float(norm[key].strip().replace("h","").replace("H",""))
                task["estimatedHours"] = val
            except ValueError:
                pass
            break
    return task


# ─── Request Handler ──────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f"  {self.address_string()} — {fmt % args}")

    def get_session(self):
        cookies = SimpleCookie(self.headers.get("Cookie", ""))
        if "session" in cookies:
            return verify_session(cookies["session"].value)
        return None

    def session_cookie(self, user):
        token = create_session(user)
        host = self.headers.get("Host", "")
        is_local = host.startswith("localhost") or host.startswith("127.")
        secure = "" if is_local else "Secure; "
        return f"session={token}; HttpOnly; {secure}SameSite=Lax; Path=/; Max-Age=2592000"

    def ok(self, content_type, body):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", len(body))
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "SAMEORIGIN")
        self.send_header("Referrer-Policy", "strict-origin-when-cross-origin")
        if content_type == "application/json":
            self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def redirect(self, location, extra=None):
        self.send_response(302)
        self.send_header("Location", location)
        if extra:
            for k, v in extra.items():
                self.send_header(k, v)
        self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        qs   = parse_qs(urlparse(self.path).query)

        # Static files
        if path == "/manifest.json":
            return self.ok("application/manifest+json", MANIFEST)
        if path == "/sw.js":
            return self.ok("application/javascript", SERVICE_WORKER)
        if path.startswith("/static/"):
            fp = os.path.join(STATIC_DIR, os.path.basename(path))
            if os.path.exists(fp):
                mime = mimetypes.guess_type(fp)[0] or "application/octet-stream"
                with open(fp, "rb") as f:
                    return self.ok(mime, f.read())
            self.send_response(404); self.end_headers(); return

        # Login
        if path == "/login":
            return self.ok("text/html; charset=utf-8", load_template("login.html").encode())

        # OAuth
        if path == "/auth/google":
            if not GOOGLE_CLIENT_ID:
                return self.ok("text/plain", b"GOOGLE_CLIENT_ID not configured")
            redirect_uri = request_redirect_uri(self)
            return self.redirect(google_auth_url(create_state(), redirect_uri))

        if path == "/auth/callback":
            code  = qs.get("code",  [None])[0]
            state_tok = qs.get("state", [None])[0]
            if not code or not state_tok or not verify_state(state_tok):
                return self.redirect("/login")
            try:
                redirect_uri = request_redirect_uri(self)
                tokens    = exchange_code(code, redirect_uri)
                user_info = get_google_user(tokens["access_token"])
                user = {
                    "id":      user_info["sub"],
                    "name":    user_info.get("name", ""),
                    "email":   user_info.get("email", ""),
                    "picture": user_info.get("picture", ""),
                }
                # Check for pending invite cookie
                cookies = SimpleCookie(self.headers.get("Cookie",""))
                invite_token = cookies["pending_invite"].value if "pending_invite" in cookies else None
                joined_group = None
                if invite_token and db is not None:
                    joined_group = db["groups"].find_one({"members.inviteToken": invite_token})
                    if joined_group:
                        for m in joined_group["members"]:
                            if m.get("inviteToken") == invite_token and m.get("status") == "pending":
                                m["userId"] = user["id"]
                                m["name"] = user.get("name","")
                                m["picture"] = user.get("picture","")
                                m["status"] = "active"
                                m["joinedAt"] = int(time.time() * 1000)
                                m.pop("inviteToken", None)
                                break
                        save_group(joined_group, user["id"])
                # Determine redirect destination after login
                redirect_to = "/mytasks"
                if joined_group and joined_group.get("slug"):
                    redirect_to = f"/groups/{joined_group['slug']}"
                elif "redirect_after_login" in cookies:
                    rval = cookies["redirect_after_login"].value
                    if rval and rval.startswith("/") and "//" not in rval and "@" not in rval:
                        redirect_to = rval
                # Send response manually to support clearing cookies
                self.send_response(302)
                self.send_header("Location", redirect_to)
                self.send_header("Set-Cookie", self.session_cookie(user))
                if invite_token:
                    self.send_header("Set-Cookie", "pending_invite=; Path=/; Max-Age=0")
                self.send_header("Set-Cookie", "redirect_after_login=; Path=/; Max-Age=0")
                self.end_headers()
                return
            except Exception as e:
                print(f"OAuth error: {e}")
                return self.redirect("/login")

        # API — require auth
        if path == "/api/health":
            health = {"status": "ok", "database": DB_STATUS}
            if DB_STATUS == "local_files":
                health["warning"] = "Data will be lost on deploy. Set MONGODB_URI to use MongoDB."
            return self.ok("application/json", json.dumps(health).encode())

        # Groups
        if path == "/api/groups":
            user = self.get_session()
            if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
            groups = load_groups_for_user(user["id"])
            result = []
            for g in groups:
                result.append({
                    "_id": g["_id"], "name": g["name"],
                    "slug": g.get("slug",""),
                    "description": g.get("description",""),
                    "color": g.get("color","#3B82F6"),
                    "createdBy": g.get("createdBy"),
                    "members": g.get("members",[]),
                    "taskCount": len(g.get("tasks",[])),
                    "doneCount": sum(1 for t in g.get("tasks",[]) if t.get("done")),
                })
            return self.ok("application/json", json.dumps(result).encode())

        if path.startswith("/api/groups/"):
            parts = path.split("/")
            # GET /api/groups/by-slug/{slug}
            if len(parts) == 5 and parts[3] == "by-slug":
                slug = parts[4]
                user = self.get_session()
                if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
                if db is None: return self.ok("application/json", b'{"error":"not found"}')
                g = db["groups"].find_one({"slug": slug})
                if not g: return self.ok("application/json", b'{"error":"not found"}')
                role = get_member_role(g, user["id"])
                if not role: return self.ok("application/json", b'{"error":"forbidden"}')
                out = dict(g); out["myRole"] = role
                return self.ok("application/json", json.dumps(out).encode())

            # GET /api/groups/{id}/stream  — Server-Sent Events for real-time sync
            if len(parts) == 5 and parts[4] == "stream":
                group_id = parts[3]
                user = self.get_session()
                if not user:
                    self.send_response(401); self.end_headers(); return
                g = load_group(group_id)
                if not g:
                    self.send_response(404); self.end_headers(); return
                if not get_member_role(g, user["id"]):
                    self.send_response(403); self.end_headers(); return
                client_q = queue.Queue()
                with _sse_lock:
                    _sse_clients.setdefault(group_id, []).append(client_q)
                try:
                    self.send_response(200)
                    self.send_header("Content-Type", "text/event-stream")
                    self.send_header("Cache-Control", "no-cache")
                    self.send_header("Connection", "keep-alive")
                    self.send_header("X-Accel-Buffering", "no")
                    self.end_headers()
                    self.wfile.write(b"data: {\"type\":\"connected\"}\n\n")
                    self.wfile.flush()
                    while True:
                        try:
                            msg = client_q.get(timeout=25)
                            self.wfile.write(msg)
                            self.wfile.flush()
                        except queue.Empty:
                            # heartbeat keeps proxies/load-balancers from closing the connection
                            self.wfile.write(b": ping\n\n")
                            self.wfile.flush()
                except (BrokenPipeError, ConnectionResetError, OSError):
                    pass
                finally:
                    with _sse_lock:
                        bucket = _sse_clients.get(group_id, [])
                        try: bucket.remove(client_q)
                        except ValueError: pass
                return

            # GET /api/groups/{id}/activities
            if len(parts) == 5 and parts[4] == "activities":
                group_id = parts[3]
                user = self.get_session()
                if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
                g = load_group(group_id)
                if not g: return self.ok("application/json", b'{"error":"not found"}')
                role = get_member_role(g, user["id"])
                if not role: return self.ok("application/json", b'{"error":"forbidden"}')
                activities = list(reversed(g.get("activities", [])))[:20]
                return self.ok("application/json", json.dumps(activities).encode())

            # /api/groups/{id}
            if len(parts) == 4:
                group_id = parts[3]
                user = self.get_session()
                if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
                g = load_group(group_id)
                if not g: return self.ok("application/json", b'{"error":"not found"}')
                role = get_member_role(g, user["id"])
                if not role: return self.ok("application/json", b'{"error":"forbidden"}')
                out = dict(g); out["_id"] = g["_id"]; out["slug"] = g.get("slug",""); out["myRole"] = role
                return self.ok("application/json", json.dumps(out).encode())

        # Notifications
        if path == "/api/notifications":
            user = self.get_session()
            if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
            if db is None: return self.ok("application/json", b'[]')
            notifs = list(db["notifications"].find({"userId": user["id"]}).sort("createdAt", -1).limit(50))
            for n in notifs:
                n.pop("_id_obj", None)
                n["_id"] = n.get("_id","")
            return self.ok("application/json", json.dumps(notifs).encode())

        # Invite token handler
        if path.startswith("/invite/"):
            token = path[8:]
            user = self.get_session()
            if db is None:
                return self.ok("text/html; charset=utf-8", b"<p>Database not available</p>")
            g = db["groups"].find_one({"members.inviteToken": token})
            if not g:
                return self.ok("text/html; charset=utf-8", b"<p>Invalid or expired invite link.</p>")
            if not user:
                # Store token in cookie, redirect to login
                self.send_response(302)
                self.send_header("Location", "/login")
                host = self.headers.get("Host", "")
                is_local = host.startswith("localhost") or host.startswith("127.")
                secure_flag = "" if is_local else "Secure; "
                self.send_header("Set-Cookie", f"pending_invite={token}; HttpOnly; {secure_flag}SameSite=Lax; Path=/; Max-Age=600")
                self.end_headers()
                return
            # Join the group
            for m in g["members"]:
                if m.get("inviteToken") == token and m.get("status") == "pending":
                    m["userId"] = user["id"]
                    m["name"] = user.get("name","")
                    m["picture"] = user.get("picture","")
                    m["status"] = "active"
                    m["joinedAt"] = int(time.time() * 1000)
                    m.pop("inviteToken", None)
                    break
            save_group(g, user["id"])
            create_notification(g["createdBy"], "member_joined", g["_id"], g["name"], from_user=user.get("name",""))
            record_activity(g["_id"], {
                "type": "member_joined",
                "userId": user["id"],
                "userName": user.get("name", ""),
                "timestamp": int(time.time() * 1000),
            })
            # Email the group admin (creator) that someone joined
            if smtp_configured() and g.get("createdBy"):
                admin_member = next((m for m in g["members"] if m.get("userId") == g["createdBy"]), None)
                if admin_member and admin_member.get("email") and admin_member["email"] != user.get("email",""):
                    host = self.headers.get("Host","")
                    is_local = host.startswith("localhost") or host.startswith("127.")
                    app_url = f"{'http' if is_local else 'https'}://{host}/"
                    html = build_member_joined_email(
                        member_name=user.get("name","Someone"),
                        member_picture=user.get("picture",""),
                        group_name=g["name"],
                        app_url=app_url,
                    )
                    _ok, _err = send_email(
                        to_email=admin_member["email"],
                        subject=f'{user.get("name","Someone")} joined your group "{g["name"]}"',
                        html_body=html,
                    )
                    if not _ok:
                        print(f"[member-joined-email] failed: {_err}")
            # Redirect to the group board
            group_slug = g.get("slug","")
            redirect_target = f"/groups/{group_slug}" if group_slug else "/groups"
            return self.redirect(redirect_target)

        if path == "/api/settings":
            user = self.get_session()
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            data = json.dumps(load_settings(user["id"])).encode()
            return self.ok("application/json", data)

        if path == "/api/todos":
            user = self.get_session()
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            data = json.dumps(load_todos(user["id"])).encode()
            return self.ok("application/json", data)

        if path == "/api/test-email":
            user = self.get_session()
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            status = smtp_status()
            if not smtp_configured():
                missing = [k for k, v in {"SMTP_HOST": SMTP_HOST, "SMTP_USER": SMTP_USER, "SMTP_PASS": SMTP_PASS}.items() if not v]
                status["error"] = f"Missing env vars: {', '.join(missing)}"
                return self.ok("application/json", json.dumps(status, indent=2).encode())
            success, error_msg = send_email(
                user["email"],
                "MyTasks Email Test",
                build_email("Email Test", "<p style='text-align:center;font-size:15px;color:#6B6B6B;'>Email is working! Sent from <strong>noreply@mytasks.bar</strong> via Brevo SMTP.</p>"),
            )
            status["success"] = success
            status["sent_to"] = user["email"]
            if not success:
                status["error"] = error_msg
            return self.ok("application/json", json.dumps(status, indent=2).encode())

        # Root redirect
        if path == "/":
            user = self.get_session()
            if not user:
                return self.redirect("/login")
            return self.redirect("/mytasks")

        # SPA app routes — serve app.html (frontend handles routing)
        SPA_ROUTES = {"/mytasks", "/analytics", "/groups"}
        if path in SPA_ROUTES or path.startswith("/groups/"):
            user = self.get_session()
            if not user:
                # Validate redirect target (must be a safe relative path)
                safe_path = path if path.startswith("/") and "//" not in path and "@" not in path else "/mytasks"
                self.send_response(302)
                self.send_header("Location", "/login")
                self.send_header("Set-Cookie", f"redirect_after_login={safe_path}; Path=/; Max-Age=300; HttpOnly")
                self.end_headers()
                return
            return self.ok("text/html; charset=utf-8", render_app(user))

        # 404 for unknown routes
        self.send_response(404)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write(b"404 Not Found")

    def do_POST(self):
        path = urlparse(self.path).path

        if path == "/auth/logout":
            self.send_response(302)
            self.send_header("Location", "/login")
            self.send_header("Set-Cookie", "session=; HttpOnly; Path=/; Max-Age=0")
            self.end_headers()
            return

        user = self.get_session()

        if path == "/api/check-overdue":
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            try:
                check_and_send_overdue_emails(user)
            except Exception as _e:
                pass
            return self.ok("application/json", b'{"ok":true}')

        if path == "/api/todos":
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            length = int(self.headers.get("Content-Length", 0))
            body   = self.rfile.read(length)
            save_todos(user["id"], json.loads(body))
            return self.ok("application/json", b'{"ok":true}')

        if path == "/api/settings":
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            length = int(self.headers.get("Content-Length", 0))
            body   = json.loads(self.rfile.read(length))
            save_settings(user["id"], body)
            return self.ok("application/json", b'{"ok":true}')

        if path in ("/api/ai/suggest", "/api/ai/summary", "/api/ai/split"):
            if not user:
                return self.ok("application/json", b'{"error":"unauthorized"}')
            length = int(self.headers.get("Content-Length", 0))
            body   = json.loads(self.rfile.read(length))
            try:
                if path == "/api/ai/suggest":
                    result = ai_suggest_next(body.get("todos", []))
                elif path == "/api/ai/summary":
                    result = ai_daily_summary(body.get("todos", []))
                else:
                    result = ai_split_task(body.get("title", ""), body.get("description", ""))
                return self.ok("application/json", json.dumps(result).encode())
            except Exception as e:
                return self.ok("application/json", json.dumps({"error": str(e)}).encode())

        # Groups POST
        if path == "/api/groups":
            if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))
            group_name = body.get("name","").strip()
            # Generate unique slug for the new group
            existing_slugs = [g.get("slug","") for g in db["groups"].find({}, {"slug": 1})] if db is not None else []
            group_slug = generate_slug(group_name, existing_slugs)
            group = {
                "_id": new_id(),
                "name": group_name,
                "slug": group_slug,
                "description": body.get("description","").strip(),
                "color": body.get("color","#3B82F6"),
                "createdBy": user["id"],
                "createdAt": int(time.time() * 1000),
                "members": [{
                    "userId": user["id"],
                    "email": user.get("email",""),
                    "name": user.get("name",""),
                    "picture": user.get("picture",""),
                    "role": "admin",
                    "status": "active",
                    "joinedAt": int(time.time() * 1000),
                }],
                "tasks": [],
            }
            if db is not None:
                db["groups"].insert_one(group)
            return self.ok("application/json", json.dumps({"_id": group["_id"], "slug": group_slug}).encode())

        if path.startswith("/api/groups/"):
            parts = path.split("/")
            if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length)) if length else {}

            # POST /api/groups/{id}/tasks
            if len(parts) == 5 and parts[4] == "tasks":
                group_id = parts[3]
                g = load_group(group_id)
                if not g: return self.ok("application/json", b'{"error":"not found"}')
                role = get_member_role(g, user["id"])
                if not role: return self.ok("application/json", b'{"error":"forbidden"}')
                assigned_to = body.get("assignedTo", user["id"])
                if role == "member": assigned_to = user["id"]
                task = {
                    "id": new_id(),
                    "text": body.get("text","").strip(),
                    "description": "",
                    "done": False,
                    "status": "todo",
                    "priority": body.get("priority","low"),
                    "createdAt": int(time.time() * 1000),
                    "completedAt": None,
                    "startedAt": None,
                    "estimatedHours": body.get("estimatedHours", None),
                    "dueDate": body.get("dueDate", None),
                    "overdueNotified": False,
                    "createdBy": user["id"],
                    "assignedTo": assigned_to,
                    "subtasks": [],
                    "notifyOnDone": body.get("notifyOnDone", False),
                }
                g["tasks"].insert(0, task)
                save_group(g, user["id"])
                if assigned_to and assigned_to != user["id"]:
                    create_notification(assigned_to, "task_assigned", group_id, g["name"],
                        task_text=task["text"], from_user=user.get("name",""))
                    # Send email to assignee if SMTP configured
                    if smtp_configured():
                        assignee_member = next((m for m in g["members"] if m.get("userId") == assigned_to), None)
                        if assignee_member and assignee_member.get("email"):
                            host = self.headers.get("Host","")
                            is_local = host.startswith("localhost") or host.startswith("127.")
                            app_url = f"{'http' if is_local else 'https'}://{host}/"
                            html = build_task_assigned_email(
                                assigner_name=user.get("name","Someone"),
                                task_text=task["text"],
                                group_name=g["name"],
                                group_color=g.get("color","#2563EB"),
                                app_url=app_url,
                            )
                            _ok, _err = send_email(
                                to_email=assignee_member["email"],
                                subject=f'New task: "{task["text"]}" assigned to you in {g["name"]}',
                                html_body=html,
                            )
                            if not _ok:
                                print(f"[task-assigned-email] failed: {_err}")
                # Record activity
                record_activity(g["_id"], {
                    "type": "task_created",
                    "userId": user["id"],
                    "userName": user.get("name", ""),
                    "taskText": task["text"],
                    "timestamp": int(time.time() * 1000),
                })
                return self.ok("application/json", json.dumps(task).encode())

            # POST /api/groups/{id}/invite
            if len(parts) == 5 and parts[4] == "invite":
                group_id = parts[3]
                g = load_group(group_id)
                if not g: return self.ok("application/json", b'{"error":"not found"}')
                role = get_member_role(g, user["id"])
                if role != "admin": return self.ok("application/json", b'{"error":"forbidden"}')
                email = body.get("email","").strip().lower()
                invite_role = body.get("role","member")
                token = new_id()
                for m in g["members"]:
                    if m.get("email","").lower() == email:
                        return self.ok("application/json", b'{"error":"already invited"}')
                g["members"].append({
                    "userId": None,
                    "email": email,
                    "name": email.split("@")[0],
                    "picture": None,
                    "role": invite_role,
                    "status": "pending",
                    "joinedAt": None,
                    "inviteToken": token,
                })
                save_group(g, user["id"])
                host = self.headers.get("Host","")
                is_local = host.startswith("localhost") or host.startswith("127.")
                scheme = "http" if is_local else "https"
                invite_url = f"{scheme}://{host}/invite/{token}"
                email_sent = False
                smtp_ok = smtp_configured()
                print(f"[invite] smtp_configured={smtp_ok}, target={email}, group={g['name']}")
                if smtp_ok:
                    html = build_invite_email(
                        inviter_name=user.get("name", "Someone"),
                        inviter_picture=user.get("picture", ""),
                        group_name=g["name"],
                        group_color=g.get("color", "#2563EB"),
                        invite_url=invite_url,
                    )
                    email_sent, email_error = send_email(
                        to_email=email,
                        subject=f'{user.get("name","Someone")} invited you to "{g["name"]}" on MyTasks',
                        html_body=html,
                    )
                    if not email_sent:
                        print(f"[invite] email failed: {email_error}")
                else:
                    email_error = "SMTP not configured"
                    print(f"[invite] SMTP not configured — SMTP_HOST={repr(SMTP_HOST)} SMTP_USER={repr(SMTP_USER)} SMTP_PASS={'set' if SMTP_PASS else 'NOT SET'}")
                resp = {"token": token, "inviteUrl": invite_url,
                        "emailSent": email_sent, "smtpConfigured": smtp_ok}
                if not email_sent:
                    resp["emailError"] = email_error
                return self.ok("application/json", json.dumps(resp).encode())

        # POST /api/groups/{id}/import — multipart CSV/XLSX upload
        if path.startswith("/api/groups/") and path.endswith("/import"):
            parts_imp = path.split("/")
            if len(parts_imp) == 5 and parts_imp[4] == "import":
                if not user: return self.ok("application/json", b'{"error":"unauthorized"}')
                group_id = parts_imp[3]
                g = load_group(group_id)
                if not g: return self.ok("application/json", b'{"error":"not found"}')
                role = get_member_role(g, user["id"])
                if role not in ("admin", "manager"):
                    return self.ok("application/json", b'{"error":"forbidden"}')
                length = int(self.headers.get("Content-Length", 0))
                body = json.loads(self.rfile.read(length)) if length else {}
                rows = []
                filename = body.get("filename", "")
                try:
                    file_data = base64.b64decode(body.get("data", ""))
                    if filename.lower().endswith(".csv"):
                        rows = parse_csv_rows(file_data)
                    elif filename.lower().endswith((".xlsx", ".xls")):
                        rows = parse_xlsx_rows(file_data)
                    else:
                        return self.ok("application/json", json.dumps({"error": "unsupported_file"}).encode())
                except Exception as e:
                    print(f"[import] parse error: {e}")
                    return self.ok("application/json", json.dumps({"error": str(e)}).encode())
                if not rows:
                    return self.ok("application/json", json.dumps({"error": "no_tasks"}).encode())
                headers_list = list(rows[0].keys()) if rows else []
                platform = detect_platform(headers_list)
                tasks = []
                assignee_emails = set()
                assignee_names = set()
                for row in rows[:2000]:  # cap at 2000 rows
                    t = map_row_to_task(row, platform)
                    if t:
                        tasks.append(t)
                        if t.get("assigneeEmail"):
                            assignee_emails.add(t["assigneeEmail"])
                        elif t.get("assigneeName"):
                            assignee_names.add(t["assigneeName"])
                # Find unmatched assignees (emails not in members)
                member_emails = {m.get("email","").lower() for m in g["members"] if m.get("email")}
                unmatched = [e for e in assignee_emails if e not in member_emails]
                # Add tasks to group
                g["tasks"] = tasks + g.get("tasks", [])
                save_group(g, user["id"])
                record_activity(g["_id"], {
                    "type": "task_created",
                    "userId": user["id"],
                    "userName": user.get("name", ""),
                    "taskText": f"Imported {len(tasks)} tasks from {platform}",
                    "timestamp": int(time.time() * 1000),
                })
                counts = {
                    "todo": sum(1 for t in tasks if t["status"] == "todo"),
                    "in_progress": sum(1 for t in tasks if t["status"] == "in_progress"),
                    "done": sum(1 for t in tasks if t["status"] == "done"),
                }
                return self.ok("application/json", json.dumps({
                    "ok": True,
                    "count": len(tasks),
                    "platform": platform,
                    "counts": counts,
                    "unmatchedAssignees": unmatched,
                }).encode())

        # POST /api/notifications/read
        if path == "/api/notifications/read":
            if not user or db is None: return self.ok("application/json", b'{"ok":true}')
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length)) if length else {}
            if body.get("all"):
                db["notifications"].update_many({"userId": user["id"]}, {"$set": {"read": True}})
            elif body.get("ids"):
                db["notifications"].update_many({"_id": {"$in": body["ids"]}}, {"$set": {"read": True}})
            return self.ok("application/json", b'{"ok":true}')

        self.send_response(404); self.end_headers()

    def do_PUT(self):
        path = urlparse(self.path).path
        parts = path.split("/")
        user = self.get_session()
        if not user:
            self.ok("application/json", b'{"error":"unauthorized"}'); return
        length = int(self.headers.get("Content-Length", 0))
        body = json.loads(self.rfile.read(length)) if length else {}

        # PUT /api/groups/{id}
        if len(parts) == 4 and parts[1]=="api" and parts[2]=="groups":
            g = load_group(parts[3])
            if not g: return self.ok("application/json", b'{"error":"not found"}')
            if get_member_role(g, user["id"]) != "admin":
                return self.ok("application/json", b'{"error":"forbidden"}')
            if "name" in body: g["name"] = body["name"]
            if "description" in body: g["description"] = body["description"]
            if "color" in body: g["color"] = body["color"]
            save_group(g, user["id"])
            return self.ok("application/json", b'{"ok":true}')

        # PUT /api/groups/{id}/tasks/{taskId}
        if len(parts) == 6 and parts[1]=="api" and parts[2]=="groups" and parts[4]=="tasks":
            g = load_group(parts[3])
            if not g: return self.ok("application/json", b'{"error":"not found"}')
            role = get_member_role(g, user["id"])
            if not role: return self.ok("application/json", b'{"error":"forbidden"}')
            activity_to_record = None
            for t in g["tasks"]:
                if t["id"] == parts[5]:
                    if "text" in body: t["text"] = body["text"]
                    if "description" in body: t["description"] = body["description"]
                    if "priority" in body: t["priority"] = body["priority"]
                    if "subtasks" in body: t["subtasks"] = body["subtasks"]
                    if "estimatedHours" in body: t["estimatedHours"] = body["estimatedHours"]
                    if "dueDate" in body: t["dueDate"] = body["dueDate"]
                    if "startedAt" in body: t["startedAt"] = body["startedAt"]
                    if "overdueNotified" in body: t["overdueNotified"] = body["overdueNotified"]
                    if "notifyOnDone" in body: t["notifyOnDone"] = body["notifyOnDone"]
                    if "status" in body:
                        old_status = t.get("status", "todo")
                        t["status"] = body["status"]
                        if body["status"] == "done":
                            t["done"] = True
                            if not t.get("completedAt"):
                                t["completedAt"] = int(time.time() * 1000)
                            if old_status != "done":
                                activity_to_record = {"type": "task_completed", "taskText": t["text"]}
                        elif body["status"] == "in_progress":
                            t["done"] = False
                            t["completedAt"] = None
                            if not t.get("startedAt"):
                                t["startedAt"] = int(time.time() * 1000)
                            if old_status != "in_progress":
                                activity_to_record = {"type": "task_started", "taskText": t["text"]}
                        elif body["status"] == "todo":
                            t["done"] = False
                            t["completedAt"] = None
                            t["startedAt"] = None
                    if "done" in body:
                        t["done"] = body["done"]
                        if body["done"]:
                            t["completedAt"] = int(time.time() * 1000)
                            t["status"] = "done"
                            if not activity_to_record:
                                activity_to_record = {"type": "task_completed", "taskText": t["text"]}
                        else:
                            t["completedAt"] = None
                            if t.get("status") == "done":
                                t["status"] = "todo"
                        if body["done"]:
                            for m in g["members"]:
                                if m.get("status")=="active" and m.get("userId") != user["id"]:
                                    create_notification(m["userId"],"task_completed",g["_id"],g["name"],
                                        task_text=t["text"],from_user=user.get("name",""))
                    if "assignedTo" in body and role in ("admin","manager"):
                        old = t.get("assignedTo")
                        t["assignedTo"] = body["assignedTo"]
                        if body["assignedTo"] and body["assignedTo"] != user["id"] and body["assignedTo"] != old:
                            create_notification(body["assignedTo"],"task_assigned",g["_id"],g["name"],
                                task_text=t["text"],from_user=user.get("name",""))
                            # Find target user name
                            target_member = next((m for m in g["members"] if m.get("userId") == body["assignedTo"]), None)
                            if not activity_to_record:
                                activity_to_record = {"type": "task_assigned", "taskText": t["text"],
                                    "targetUser": target_member.get("name","") if target_member else ""}
                    break
            save_group(g, user["id"])
            if activity_to_record:
                activity_to_record.update({
                    "userId": user["id"],
                    "userName": user.get("name", ""),
                    "timestamp": int(time.time() * 1000),
                })
                record_activity(g["_id"], activity_to_record)
            # Send "notify when done" email if task was just completed
            if activity_to_record and activity_to_record.get("type") == "task_completed" and smtp_configured():
                completed_task = next((t for t in g["tasks"] if t["id"] == parts[5]), None)
                if completed_task and completed_task.get("notifyOnDone"):
                    creator_id = completed_task.get("createdBy")
                    if creator_id:
                        creator_member = next((m for m in g["members"] if m.get("userId") == creator_id), None)
                        if creator_member and creator_member.get("email"):
                            html = build_task_done_email(
                                completer_name=user.get("name","Someone"),
                                task_text=completed_task["text"],
                                group_name=g["name"],
                                group_color=g.get("color","#22C55E"),
                            )
                            subj = f'✅ Done: "{completed_task["text"]}" in {g["name"]}'
                            _ok2, _err2 = send_email(creator_member["email"], subj, html)
                            if not _ok2:
                                print(f"[notify-done-email] failed: {_err2}")
            return self.ok("application/json", b'{"ok":true}')

        # PUT /api/groups/{id}/members/{userId}
        if len(parts) == 6 and parts[1]=="api" and parts[2]=="groups" and parts[4]=="members":
            g = load_group(parts[3])
            if not g: return self.ok("application/json", b'{"error":"not found"}')
            if get_member_role(g, user["id"]) != "admin":
                return self.ok("application/json", b'{"error":"forbidden"}')
            target_id = parts[5]
            for m in g["members"]:
                if m.get("userId") == target_id:
                    m["role"] = body.get("role", m["role"])
                    break
            save_group(g, user["id"])
            return self.ok("application/json", b'{"ok":true}')

        self.send_response(404); self.end_headers()

    def do_DELETE(self):
        path = urlparse(self.path).path
        parts = path.split("/")
        user = self.get_session()
        if not user:
            self.ok("application/json", b'{"error":"unauthorized"}'); return

        # DELETE /api/groups/{id}
        if len(parts) == 4 and parts[1]=="api" and parts[2]=="groups":
            g = load_group(parts[3])
            if not g: return self.ok("application/json", b'{"error":"not found"}')
            if get_member_role(g, user["id"]) != "admin":
                return self.ok("application/json", b'{"error":"forbidden"}')
            if db is not None: db["groups"].delete_one({"_id": parts[3]})
            return self.ok("application/json", b'{"ok":true}')

        # DELETE /api/groups/{id}/tasks/{taskId}
        if len(parts) == 6 and parts[1]=="api" and parts[2]=="groups" and parts[4]=="tasks":
            g = load_group(parts[3])
            if not g: return self.ok("application/json", b'{"error":"not found"}')
            role = get_member_role(g, user["id"])
            if not role: return self.ok("application/json", b'{"error":"forbidden"}')
            task = next((t for t in g["tasks"] if t["id"]==parts[5]), None)
            if task:
                if role in ("admin","manager") or task.get("createdBy")==user["id"]:
                    g["tasks"] = [t for t in g["tasks"] if t["id"]!=parts[5]]
            save_group(g, user["id"])
            return self.ok("application/json", b'{"ok":true}')

        # DELETE /api/groups/{id}/members/{userId|email}
        if len(parts) == 6 and parts[1]=="api" and parts[2]=="groups" and parts[4]=="members":
            g = load_group(parts[3])
            if not g: return self.ok("application/json", b'{"error":"not found"}')
            role = get_member_role(g, user["id"])
            target_id = parts[5]
            # Allow self-removal or admin removing anyone
            if target_id != user["id"] and role != "admin":
                return self.ok("application/json", b'{"error":"forbidden"}')
            # Match by userId (active members) OR by email (pending members, userId is null)
            g["members"] = [m for m in g["members"]
                if m.get("userId") != target_id
                and m.get("email","").lower() != target_id.lower()]
            for t in g["tasks"]:
                if t.get("assignedTo") == target_id: t["assignedTo"] = None
            save_group(g, user["id"])
            print(f"[remove-member] removed target={target_id} from group={g['_id']}")
            return self.ok("application/json", b'{"ok":true}')

        self.send_response(404); self.end_headers()


if __name__ == "__main__":
    # Migrate existing groups to have slugs
    if db is not None:
        migrate_group_slugs()
    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print(f"✓ Running at http://localhost:{PORT}")
    if not GOOGLE_CLIENT_ID:   print("⚠  GOOGLE_CLIENT_ID not set")
    if not ANTHROPIC_API_KEY:  print("⚠  ANTHROPIC_API_KEY not set — AI features disabled")
    print_smtp_status()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
